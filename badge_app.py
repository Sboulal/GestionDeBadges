"""
Badge Management API Server with Brother QL Label Printer Integration
Flask REST API for badge management and label printing - Windows Compatible
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import sqlite3
from datetime import datetime
import json
import qrcode
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import os
import sys
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Brother QL Label Printer imports
from PIL import Image, ImageDraw, ImageFont
from brother_ql.conversion import convert
from brother_ql.backends.helpers import send
from brother_ql.raster import BrotherQLRaster
import subprocess

app = Flask(__name__)
CORS(app)  # Enable CORS for cross-origin requests

DB_NAME = 'badges.db'
EXTERNAL_API_URL = 'http://badges.eevent.ma/api/getbadges'

# Brother QL Printer Configuration
PRINTER_MODEL = "QL-810W"

# Determine OS and set appropriate backend
IS_WINDOWS = sys.platform.startswith('win')
PRINTER_BACKEND = "pyusb" if not IS_WINDOWS else "network"  # or "win_usb"

# Windows-compatible printer identifier
# You'll need to update this based on your printer connection:
# For USB on Windows: use "usb://0x04f9:0x209c" or the printer name
# For Network: use "tcp://192.168.1.100" (replace with your printer's IP)
if IS_WINDOWS:
    # Option 1: Network printing (recommended for Windows)
    PRINTER_IDENTIFIER = "tcp://192.168.1.100"  # CHANGE THIS to your printer's IP
    
    # Option 2: USB printing on Windows (uncomment if using USB)
    # PRINTER_IDENTIFIER = "usb://0x04f9:0x209c"
    # PRINTER_BACKEND = "pyusb"  # or try "win_usb" if pyusb doesn't work
else:
    PRINTER_IDENTIFIER = "usb://0x04f9:0x209c"

# ==================== Database Helper Functions ====================

def get_db_connection():
    """Create database connection"""
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """Initialize database"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nom TEXT NOT NULL,
            prenom TEXT NOT NULL,
            valide INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS print_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            printed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')
    
    conn.commit()
    conn.close()

# ==================== Windows Font Helper Functions ====================

def get_font_path(font_name, font_size):
    """Get font path based on operating system"""
    font = None
    
    if IS_WINDOWS:
        # Windows font paths
        windows_fonts_dir = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts')
        font_paths = [
            os.path.join(windows_fonts_dir, 'arial.ttf'),
            os.path.join(windows_fonts_dir, 'arialbd.ttf'),
            os.path.join(windows_fonts_dir, 'calibri.ttf'),
            os.path.join(windows_fonts_dir, 'calibrib.ttf'),
            os.path.join(windows_fonts_dir, 'segoeui.ttf'),
            os.path.join(windows_fonts_dir, 'segoeuib.ttf'),
        ]
    else:
        # Linux/Unix font paths
        font_paths = [
            'arial.ttf',
            'calibri.ttf',
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
            '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
            'DejaVuSans.ttf',
        ]
    
    # Try each font path
    for font_path in font_paths:
        if os.path.exists(font_path):
            try:
                font = ImageFont.truetype(font_path, font_size)
                print(f"Using font: {font_path}")
                return font
            except Exception as e:
                print(f"Could not load font {font_path}: {e}")
                continue
    
    # Fallback to default font
    print("Warning: No TrueType font found, using default font")
    return ImageFont.load_default()

# ==================== Brother QL Label Printer Functions ====================

def check_supported_models():
    """Check supported models by running brother_ql info models."""
    try:
        result = subprocess.run(['brother_ql', 'info', 'models'], 
                              capture_output=True, text=True, shell=IS_WINDOWS)
        return result.stdout
    except Exception as e:
        return f"Error checking models: {e}"

def create_label_image(first_name, last_name):
    """Create label image for Brother QL printer"""
    # Label dimensions for 29mm x 90mm (in pixels, as expected by brother_ql)
    label_width = 991   # Effective printable width for 90mm
    label_height = 306  # 29mm at 300 DPI

    # Create a grayscale image
    image = Image.new("L", (label_width, label_height), "white")
    draw = ImageDraw.Draw(image)

    # Combine first and last name
    full_name = f"{first_name} {last_name}"

    # Maximum dimensions for text (95% of width, 90% of height for larger text)
    max_text_width = int(label_width * 0.95)    # ≈941px
    max_text_height = int(label_height * 0.9)   # ≈275px

    # Start with a large font size and scale down
    font_size = 120  # Start large for bigger text
    font = None
    
    while font_size > 20:  # Minimum font size
        font = get_font_path("arial.ttf", font_size)
        
        if font:
            text_bbox = draw.textbbox((0, 0), full_name, font=font)
            text_width = text_bbox[2] - text_bbox[0]
            text_height = text_bbox[3] - text_bbox[1]

            # Check if text fits within max dimensions
            if text_width <= max_text_width and text_height <= max_text_height:
                break
            font_size -= 5  # Reduce font size and try again
        else:
            break

    # Use minimum font size if text still doesn't fit
    if font_size <= 20:
        font_size = 20
        font = get_font_path("arial.ttf", font_size)

    # Calculate text size and position for centering
    text_bbox = draw.textbbox((0, 0), full_name, font=font)
    text_width = text_bbox[2] - text_bbox[0]
    text_height = text_bbox[3] - text_bbox[1]

    # Center text horizontally
    x = (label_width - text_width) // 2

    # Center text vertically with adjustment for font metrics
    if hasattr(font, 'getmetrics'):
        try:
            ascent, descent = font.getmetrics()
            text_visual_height = ascent - descent
        except:
            text_visual_height = text_height
    else:
        text_visual_height = text_height
    
    y = (label_height - text_visual_height) // 2 - text_bbox[1]

    # Draw text
    draw.text((x, y), full_name, fill="black", font=font)

    print(f"Using font size: {font_size}pt for '{full_name}'")
    return image

def print_to_brother_ql(first_name, last_name, 
                       printer_identifier=PRINTER_IDENTIFIER, 
                       model=PRINTER_MODEL):
    """Print label to Brother QL printer"""
    # Check if model is valid
    supported_models = check_supported_models()
    if model not in supported_models:
        error_msg = f"Error: Model '{model}' not recognized. Supported models:\n{supported_models}"
        print(error_msg)
        return {"status": "error", "message": error_msg}

    # Create label image
    image = create_label_image(first_name, last_name)
    if image is None:
        error_msg = "Failed to create label image"
        print(error_msg)
        return {"status": "error", "message": error_msg}

    # Convert and send to printer
    try:
        qlr = BrotherQLRaster(model)
        qlr.exception_on_warning = True
        instructions = convert(
            qlr=qlr,
            images=[image],
            label="29x90",  # 29mm x 90mm label
            rotate="90",    # Rotate 90 degrees
            threshold=70.0,
            dither=False,
            compress=False,
            red=False,      # Set to True for black/red labels
            dpi_600=False,
            hq=True,
            cut=True
        )

        # Send to printer
        send(
            instructions=instructions,
            printer_identifier=printer_identifier,
            backend_identifier=PRINTER_BACKEND,
            blocking=True
        )
        print(f"Label printed successfully for '{first_name} {last_name}'!")
        return {"status": "success", "message": "Label printed successfully"}
    except Exception as e:
        error_msg = f"Error printing label: {str(e)}"
        print(error_msg)
        print(f"Printer Identifier: {printer_identifier}")
        print(f"Backend: {PRINTER_BACKEND}")
        return {"status": "error", "message": error_msg}

# ==================== QR Code Generation (for PDF fallback) ====================

def generate_qr_code(data):
    """Generate QR code image"""
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    return img

def create_badge_pdf(user_data):
    """Create badge PDF with QR code (fallback method)"""
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Title
    c.setFont("Helvetica-Bold", 24)
    c.drawCentredString(width/2, height - 100, "BADGE D'ACCÈS")
    
    # User information
    c.setFont("Helvetica", 18)
    full_name = f"{user_data.get('first_name', user_data.get('prenom', ''))} {user_data.get('last_name', user_data.get('nom', ''))}"
    c.drawCentredString(width/2, height - 150, full_name.upper())
    
    # ID
    c.setFont("Helvetica", 12)
    c.drawCentredString(width/2, height - 180, f"ID: {user_data.get('id', 'N/A')}")
    
    # Generate QR code
    qr_data = json.dumps({
        'id': user_data.get('id', 0),
        'nom': user_data.get('last_name', user_data.get('nom', '')),
        'prenom': user_data.get('first_name', user_data.get('prenom', '')),
        'timestamp': datetime.now().isoformat()
    })
    
    qr_img = generate_qr_code(qr_data)
    
    # Save QR code to buffer
    qr_buffer = BytesIO()
    qr_img.save(qr_buffer, format='PNG')
    qr_buffer.seek(0)
    
    # Add QR code to PDF
    qr_reader = ImageReader(qr_buffer)
    qr_size = 200
    c.drawImage(qr_reader, 
               (width - qr_size) / 2, 
               height - 450, 
               width=qr_size, 
               height=qr_size)
    
    # Issue date
    c.setFont("Helvetica", 10)
    c.drawCentredString(width/2, 100, 
                      f"Émis le: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    
    c.save()
    buffer.seek(0)
    return buffer

# ==================== Excel Functions ====================

def create_excel_export(badges_data):
    """Create Excel file from badges data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Badges"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['ID', 'Prénom', 'Nom', 'Validé', 'Date de création', 'Dernière modification', 'Source']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Add data
    for badge in badges_data:
        row = [
            badge.get('id', ''),
            badge.get('prenom', ''),
            badge.get('nom', ''),
            'Oui' if badge.get('valide') == 1 else 'Non',
            badge.get('created_at', ''),
            badge.get('updated_at', ''),
            badge.get('source', 'local')
        ]
        ws.append(row)
        
        # Apply border to all cells in the row
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Adjust column widths
    column_widths = [8, 20, 20, 10, 20, 20, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ==================== API Routes ====================

@app.route('/')
def index():
    """API information endpoint"""
    return jsonify({
        'name': 'Badge Management API with Brother QL Printer (Windows Compatible)',
        'version': '2.1',
        'platform': 'Windows' if IS_WINDOWS else 'Linux/Unix',
        'printer': {
            'model': PRINTER_MODEL,
            'identifier': PRINTER_IDENTIFIER,
            'backend': PRINTER_BACKEND
        },
        'endpoints': {
            'GET /api/getbadges': 'Get all badges (local + external)',
            'GET /api/getbadges/<id>': 'Get badge by ID',
            'POST /api/badges': 'Create new badge',
            'PUT /api/badges/<id>': 'Update badge',
            'DELETE /api/badges/<id>': 'Delete badge',
            'POST /print-label': 'Print badge to Brother QL printer',
            'POST /print-label-pdf': 'Generate PDF badge (fallback)',
            'GET /api/stats': 'Get statistics',
            'POST /api/validate/<id>': 'Validate badge',
            'GET /api/search': 'Search badges',
            'GET /api/export-excel': 'Export badges to Excel file'
        }
    })

@app.route('/api/getbadges', methods=['GET'])
def get_all_badges():
    """Get all badges (combines local and external data)"""
    try:
        # Get local badges
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get filter parameters
        valide = request.args.get('valide', type=int)
        search = request.args.get('search', '')
        source = request.args.get('source', 'all')
        
        local_badges = []
        if source in ['all', 'local']:
            cursor.execute('SELECT * FROM users ORDER BY id ASC')
            users = cursor.fetchall()
            
            for user in users:
                if valide is not None and user['valide'] != valide:
                    continue
                
                if search:
                    search_lower = search.lower()
                    if (search_lower not in user['nom'].lower() and 
                        search_lower not in user['prenom'].lower() and 
                        search_lower not in str(user['id'])):
                        continue
                
                local_badges.append({
                    'id': user['id'],
                    'nom': user['nom'],
                    'prenom': user['prenom'],
                    'valide': user['valide'],
                    'created_at': user['created_at'],
                    'updated_at': user['updated_at'],
                    'source': 'local'
                })
        
        conn.close()
        
        # Get external badges
        external_badges = []
        if source in ['all', 'external']:
            try:
                response = requests.get(EXTERNAL_API_URL, timeout=5)
                if response.status_code == 200:
                    external_data = response.json()
                    
                    for badge in external_data:
                        if valide is not None and badge.get('valide') != valide:
                            continue
                        
                        if search:
                            search_lower = search.lower()
                            nom = str(badge.get('nom', '')).lower()
                            prenom = str(badge.get('prenom', '')).lower()
                            badge_id = str(badge.get('id', ''))
                            
                            if search_lower not in nom and search_lower not in prenom and search_lower not in badge_id:
                                continue
                        
                        badge['source'] = 'external'
                        external_badges.append(badge)
            except requests.RequestException as e:
                print(f"Warning: Could not fetch external badges: {str(e)}")
        
        # Combine results
        all_badges = local_badges + external_badges
        all_badges.sort(key=lambda x: x.get('id', 0))  # Ascending order (oldest first)
        
        return jsonify(all_badges)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/getbadges/<int:badge_id>', methods=['GET'])
def get_badge_by_id(badge_id):
    """Get specific badge by ID"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM users WHERE id = ?', (badge_id,))
        user = cursor.fetchone()
        conn.close()
        
        if user:
            return jsonify({
                'id': user['id'],
                'nom': user['nom'],
                'prenom': user['prenom'],
                'valide': user['valide'],
                'created_at': user['created_at'],
                'updated_at': user['updated_at'],
                'source': 'local'
            })
        
        try:
            response = requests.get(f"{EXTERNAL_API_URL}/{badge_id}", timeout=5)
            if response.status_code == 200:
                badge = response.json()
                badge['source'] = 'external'
                return jsonify(badge)
        except requests.RequestException:
            pass
        
        return jsonify({'error': 'Badge not found'}), 404
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/print-label', methods=['POST'])
def print_label():
    """Print badge label to Brother QL printer"""
    try:
        data = request.get_json()
        
        # Support both formats
        nom = data.get('nom', data.get('last_name', ''))
        prenom = data.get('prenom', data.get('first_name', ''))
        user_id = data.get('id')
        
        if not nom or not prenom:
            return jsonify({'error': 'last_name and first_name are required'}), 400
        
        # If no ID provided, create new user
        if not user_id:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO users (nom, prenom, valide, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
            ''', (nom, prenom, 1, datetime.now(), datetime.now()))
            user_id = cursor.lastrowid
            conn.commit()
            conn.close()
        
        # Print to Brother QL printer
        result = print_to_brother_ql(prenom, nom)
        
        if result['status'] == 'success':
            # Log print
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO print_logs (user_id, printed_at)
                VALUES (?, ?)
            ''', (user_id, datetime.now()))
            conn.commit()
            conn.close()
            
            return jsonify({
                'status': 'success',
                'message': f'Label printed successfully for {prenom} {nom}',
                'id': user_id
            }), 200
        else:
            return jsonify(result), 500
    
    except Exception as e:
        return jsonify({'error': str(e), 'status': 'error'}), 500

@app.route('/print-label-pdf', methods=['POST'])
def print_label_pdf():
    """Generate PDF badge (fallback method when printer is not available)"""
    try:
        data = request.get_json()
        
        nom = data.get('nom', data.get('last_name', ''))
        prenom = data.get('prenom', data.get('first_name', ''))
        user_id = data.get('id')
        
        if not nom or not prenom:
            return jsonify({'error': 'last_name and first_name are required'}), 400
        
        # If no ID provided, create new user
        if not user_id:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO users (nom, prenom, valide, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
            ''', (nom, prenom, 1, datetime.now(), datetime.now()))
            user_id = cursor.lastrowid
            conn.commit()
            conn.close()
        
        # Create user data for PDF
        user_data = {
            'id': user_id,
            'nom': nom,
            'prenom': prenom,
            'last_name': nom,
            'first_name': prenom
        }
        
        # Generate PDF
        pdf_buffer = create_badge_pdf(user_data)
        
        # Log print
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO print_logs (user_id, printed_at)
            VALUES (?, ?)
        ''', (user_id, datetime.now()))
        conn.commit()
        conn.close()
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'badge_{prenom}_{nom}.pdf'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/stats', methods=['GET'])
def get_statistics():
    """Get statistics"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT COUNT(*) as total FROM users')
        total = cursor.fetchone()['total']
        
        cursor.execute('SELECT COUNT(*) as validated FROM users WHERE valide = 1')
        validated = cursor.fetchone()['validated']
        
        cursor.execute('SELECT COUNT(*) as printed FROM print_logs')
        printed = cursor.fetchone()['printed']
        
        conn.close()
        
        return jsonify({
            'total_badges': total,
            'validated_badges': validated,
            'total_prints': printed
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-excel', methods=['GET'])
def export_excel():
    """Export all badges to Excel"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM users ORDER BY id ASC')
        users = cursor.fetchall()
        conn.close()
        
        badges_data = []
        for user in users:
            badges_data.append({
                'id': user['id'],
                'nom': user['nom'],
                'prenom': user['prenom'],
                'valide': user['valide'],
                'created_at': user['created_at'],
                'updated_at': user['updated_at'],
                'source': 'local'
            })
        
        excel_buffer = create_excel_export(badges_data)
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'badges_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== Run Server ====================

if __name__ == '__main__':
    init_db()
    print("=" * 60)
    print("Badge Management API Server with Brother QL Printer")
    print("Windows Compatible Version" if IS_WINDOWS else "Linux/Unix Version")
    print("=" * 60)
    print(f"Server starting on http://127.0.0.1:5000")
    print(f"Platform: {'Windows' if IS_WINDOWS else 'Linux/Unix'}")
    print(f"Printer Model: {PRINTER_MODEL}")
    print(f"Printer Identifier: {PRINTER_IDENTIFIER}")
    print(f"Printer Backend: {PRINTER_BACKEND}")
    print(f"External API: {EXTERNAL_API_URL}")
    print("\nPrinter endpoints:")
    print("  POST /print-label       - Print to Brother QL printer")
    print("  POST /print-label-pdf   - Generate PDF (fallback)")
    if IS_WINDOWS:
        print("\n⚠️  WINDOWS SETUP NOTES:")
        print("  1. Update PRINTER_IDENTIFIER in the code with your printer's:")
        print("     - Network IP: tcp://192.168.1.100")
        print("     - OR USB: usb://0x04f9:0x209c")
        print("  2. For USB printing, you may need to install:")
        print("     - libusb: pip install libusb")
        print("     - zadig (to install WinUSB driver for the printer)")
        print("  3. Network printing is recommended for Windows")
    print("=" * 60)
    
    app.run(host='0.0.0.0', port=5000, debug=True)